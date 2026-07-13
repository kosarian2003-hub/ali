"""
Khorshid (خورشید) — Home Appliance Distribution & Repair Co.
Backend API — Flask

Responsibilities
- Reads live product data (name / price / stock) from an Excel file using openpyxl
  and serves it as JSON at GET /api/products. The file is re-read on every request
  (cheap for a catalog this size) so any edit saved in Excel is reflected on the
  very next poll — the frontend polls every 10 seconds.
- Minimal auth (signup / login / logout) backed by a JSON user store, password
  hashed with werkzeug's generator (never stored in plain text).
- A small accounting layer: every paid order becomes an order record + an invoice
  record (JSON-file "ledger"). GET /api/accounting/summary aggregates revenue,
  order count and current inventory value — the seed of a real accounting module.
- Shipping rule: a customer's first order ships free; every order after that has
  a flat shipping fee.
- POST /api/admin/upload-products lets you push a new Excel file to the server
  from any machine (see README "Local Excel + hosted site" section for why this
  endpoint exists).

Run:
    pip install flask openpyxl werkzeug
    python app.py
Then open http://localhost:5000
"""

import json
import os
import time
import uuid
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, request, session, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
FRONTEND_DIR = BASE_DIR.parent / "frontend"

PRODUCTS_XLSX = DATA_DIR / "products.xlsx"
USERS_JSON = DATA_DIR / "users.json"
ORDERS_JSON = DATA_DIR / "orders.json"
INVOICES_JSON = DATA_DIR / "invoices.json"

FREE_SHIPPING_ON_FIRST_ORDER = True
STANDARD_SHIPPING_FEE = 350_000  # تومان

app = Flask(__name__, static_folder=None)
app.secret_key = os.environ.get("KHORSHID_SECRET_KEY", "dev-secret-change-me")


# --------------------------------------------------------------------------- #
# tiny JSON "database" helpers
# --------------------------------------------------------------------------- #
def _read_json(path: Path):
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _write_json(path: Path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# --------------------------------------------------------------------------- #
# products (read live from Excel)
# --------------------------------------------------------------------------- #
def read_products_from_excel():
    if not PRODUCTS_XLSX.exists():
        return []

    wb = openpyxl.load_workbook(PRODUCTS_XLSX, data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    products = []
    for row in rows[1:]:
        if row is None or all(v is None for v in row):
            continue
        record = dict(zip(headers, row))
        try:
            products.append({
                "id": int(record.get("id")),
                "name_fa": record.get("name_fa") or "",
                "name_en": record.get("name_en") or "",
                "category_fa": record.get("category_fa") or "",
                "category_en": record.get("category_en") or "",
                "price": int(record.get("price_toman") or 0),
                "stock": int(record.get("stock") or 0),
                "image": record.get("image") or "",
                "sku": record.get("sku") or "",
            })
        except (TypeError, ValueError):
            # skip malformed rows instead of crashing the whole endpoint
            continue
    wb.close()
    return products


def decrement_stock(items):
    """
    Subtract each purchased item's qty from its stock cell in products.xlsx.
    Called right after a successful payment. Best-effort: if the file is
    locked (e.g. open in Excel) this raises, and the caller decides whether
    that should block the payment or just get logged.
    """
    if not PRODUCTS_XLSX.exists() or not items:
        return

    wb = openpyxl.load_workbook(PRODUCTS_XLSX, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    id_col = headers.index("id") + 1
    stock_col = headers.index("stock") + 1

    qty_by_id = {}
    for i in items:
        qty_by_id[int(i["id"])] = qty_by_id.get(int(i["id"]), 0) + int(i.get("qty", 1))

    for row in ws.iter_rows(min_row=2):
        cell_id = row[id_col - 1].value
        if cell_id in qty_by_id:
            current = row[stock_col - 1].value or 0
            row[stock_col - 1].value = max(0, int(current) - qty_by_id[cell_id])

    wb.save(PRODUCTS_XLSX)
    wb.close()


@app.route("/api/products", methods=["GET"])
def api_products():
    try:
        products = read_products_from_excel()
        mtime = PRODUCTS_XLSX.stat().st_mtime if PRODUCTS_XLSX.exists() else 0
        return jsonify({
            "ok": True,
            "updated_at": mtime,
            "count": len(products),
            "products": products,
        })
    except Exception as exc:  # keep the API alive even if the file is mid-save
        msg = str(exc)
        if isinstance(exc, PermissionError):
            msg = (
                "products.xlsx is locked by another program — close it in Excel "
                "(or any app that has it open) and try again."
            )
        return jsonify({"ok": False, "error": msg, "products": []}), 200


@app.route("/api/admin/upload-products", methods=["POST"])
def api_upload_products():
    """
    Push a new products.xlsx to the server.

    This is the bridge for the "site is hosted, but the Excel file lives on my
    own computer" scenario: a remote server has no way to reach into a laptop's
    filesystem on its own, so this endpoint lets a small local script (or this
    admin page) upload the file whenever it changes. Pair it with a folder-watch
    script on your machine for near-live sync — see README.
    """
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "no file provided"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "file must be .xlsx"}), 400
    f.save(PRODUCTS_XLSX)
    return jsonify({"ok": True, "message": "products.xlsx updated"})


# --------------------------------------------------------------------------- #
# auth
# --------------------------------------------------------------------------- #
@app.route("/api/auth/signup", methods=["POST"])
def signup():
    body = request.get_json(force=True, silent=True) or {}
    first_name = (body.get("first_name") or "").strip()
    last_name = (body.get("last_name") or "").strip()
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""

    if not first_name or not last_name or not email or len(password) < 6:
        return jsonify({"ok": False, "error": "invalid_input"}), 400

    users = _read_json(USERS_JSON)
    if any(u["email"] == email for u in users):
        return jsonify({"ok": False, "error": "email_exists"}), 409

    name = f"{first_name} {last_name}".strip()
    user = {
        "id": str(uuid.uuid4()),
        "first_name": first_name,
        "last_name": last_name,
        "name": name,
        "email": email,
        "password_hash": generate_password_hash(password),
        "created_at": datetime.utcnow().isoformat(),
    }
    users.append(user)
    _write_json(USERS_JSON, users)

    session["user_id"] = user["id"]
    return jsonify({"ok": True, "user": {
        "id": user["id"], "first_name": first_name, "last_name": last_name, "name": name, "email": email,
    }})


@app.route("/api/auth/login", methods=["POST"])
def login():
    body = request.get_json(force=True, silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""

    users = _read_json(USERS_JSON)
    user = next((u for u in users if u["email"] == email), None)
    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"ok": False, "error": "invalid_credentials"}), 401

    session["user_id"] = user["id"]
    return jsonify({"ok": True, "user": {
        "id": user["id"],
        "first_name": user.get("first_name", ""),
        "last_name": user.get("last_name", ""),
        "name": user["name"],
        "email": user["email"],
    }})


@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.pop("user_id", None)
    return jsonify({"ok": True})


@app.route("/api/auth/me", methods=["GET"])
def me():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"ok": True, "user": None})
    users = _read_json(USERS_JSON)
    user = next((u for u in users if u["id"] == uid), None)
    if not user:
        return jsonify({"ok": True, "user": None})
    return jsonify({"ok": True, "user": {
        "id": user["id"],
        "first_name": user.get("first_name", ""),
        "last_name": user.get("last_name", ""),
        "name": user["name"],
        "email": user["email"],
    }})


# --------------------------------------------------------------------------- #
# orders / shipping / accounting
# --------------------------------------------------------------------------- #
@app.route("/api/orders", methods=["POST"])
def create_order():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"ok": False, "error": "login_required"}), 401

    body = request.get_json(force=True, silent=True) or {}
    items = body.get("items") or []  # [{id, name, price, qty}]
    delivery = body.get("delivery") or {}  # {lat, lng, address}
    customer = body.get("customer") or {}

    if not items:
        return jsonify({"ok": False, "error": "empty_cart"}), 400

    # validate against live stock so no one can order more than what's in the warehouse
    stock_by_id = {p["id"]: p["stock"] for p in read_products_from_excel()}
    for i in items:
        pid = int(i["id"])
        requested = int(i.get("qty", 1))
        available = stock_by_id.get(pid, 0)
        if requested > available:
            return jsonify({
                "ok": False,
                "error": "insufficient_stock",
                "product_id": pid,
                "available": available,
            }), 409

    orders = _read_json(ORDERS_JSON)
    previous_orders = [o for o in orders if o.get("user_id") == uid] if uid else []

    subtotal = sum(int(i.get("price", 0)) * int(i.get("qty", 1)) for i in items)

    is_first_order = FREE_SHIPPING_ON_FIRST_ORDER and len(previous_orders) == 0
    shipping_fee = 0 if is_first_order else STANDARD_SHIPPING_FEE
    total = subtotal + shipping_fee

    order = {
        "id": str(uuid.uuid4())[:8],
        "user_id": uid,
        "customer": customer,
        "items": items,
        "delivery": delivery,
        "subtotal": subtotal,
        "shipping_fee": shipping_fee,
        "free_shipping_applied": is_first_order,
        "total": total,
        "status": "pending_payment",
        "created_at": datetime.utcnow().isoformat(),
    }
    orders.append(order)
    _write_json(ORDERS_JSON, orders)

    return jsonify({"ok": True, "order": order})


@app.route("/api/orders/<order_id>/pay", methods=["POST"])
def pay_order(order_id):
    """
    Placeholder payment confirmation. Wire this to a real gateway (e.g. Zarinpal
    / IDPay / a bank's own PSP) by calling their "request payment" API here,
    redirecting the browser to their page, and verifying the callback before
    marking the order paid. Until real credentials are supplied, this simulates
    a successful payment so the rest of the flow (invoice + accounting) works.
    """
    orders = _read_json(ORDERS_JSON)
    order = next((o for o in orders if o["id"] == order_id), None)
    if not order:
        return jsonify({"ok": False, "error": "order_not_found"}), 404

    order["status"] = "paid"
    order["paid_at"] = datetime.utcnow().isoformat()
    _write_json(ORDERS_JSON, orders)

    try:
        decrement_stock(order["items"])
    except Exception as exc:
        # the order is already paid and recorded — surface the stock-sync
        # failure separately rather than losing the sale over a locked file
        order["stock_sync_error"] = str(exc)
        _write_json(ORDERS_JSON, orders)

    invoices = _read_json(INVOICES_JSON)
    invoice = {
        "id": f"INV-{len(invoices) + 1001}",
        "order_id": order["id"],
        "customer": order["customer"],
        "subtotal": order["subtotal"],
        "shipping_fee": order["shipping_fee"],
        "total": order["total"],
        "issued_at": datetime.utcnow().isoformat(),
    }
    invoices.append(invoice)
    _write_json(INVOICES_JSON, invoices)

    return jsonify({"ok": True, "order": order, "invoice": invoice})


@app.route("/api/accounting/summary", methods=["GET"])
def accounting_summary():
    orders = _read_json(ORDERS_JSON)
    invoices = _read_json(INVOICES_JSON)
    products = read_products_from_excel()

    paid_orders = [o for o in orders if o.get("status") == "paid"]
    revenue = sum(o["total"] for o in paid_orders)
    shipping_collected = sum(o["shipping_fee"] for o in paid_orders)
    inventory_value = sum(p["price"] * p["stock"] for p in products)
    out_of_stock = [p["name_fa"] for p in products if p["stock"] == 0]

    return jsonify({
        "ok": True,
        "revenue_toman": revenue,
        "shipping_collected_toman": shipping_collected,
        "paid_orders": len(paid_orders),
        "pending_orders": len(orders) - len(paid_orders),
        "invoices_issued": len(invoices),
        "inventory_value_toman": inventory_value,
        "out_of_stock_products": out_of_stock,
    })


# --------------------------------------------------------------------------- #
# static frontend
# --------------------------------------------------------------------------- #
@app.route("/")
def serve_index():
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.route("/<path:path>")
def serve_static(path):
    return send_from_directory(FRONTEND_DIR, path)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)


# if __name__ == "__main__":
#     app.run(debug=True, port=5000)
